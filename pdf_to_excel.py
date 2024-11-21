# Import Modules  
import tabula
import jpype
import os
os.environ["JAVA_HOME"] ="/Library/Internet Plug-Ins/JavaAppletPlugin.plugin/Contents/Home"

file_path = 'alphadoku-symmetric-expert-puzzle-44'
  
# Read PDF File 
df = tabula.read_pdf(f'{file_path}.pdf')[0]

print("Converting...")
# Convert into Excel File 
df.to_excel(f'{file_path}.xlsx',index=False)


from openpyxl import load_workbook
from openpyxl.styles import Font

code = {'A': 1,
        'B': 2,
        'C': 3,
        'D': 4,
        'E': 5,
        'F': 6,
        'G': 7,
        'H': 8,
        'I': 9,
        'J': 10,
        'K': 11,
        'L': 12,
        'M': 13,
        'N': 14,
        'O': 15,
        'P': 16,
        'Q': 17,
        'R': 18,
        'S': 19,
        'T': 20,
        'U': 21,
        'V': 22,
        'W': 23,
        'X': 24,
        'Y': 25}

print("Reformatting...")
# reformat
def copy_formatting(source_wb, target_wb):
        target_ws = target_wb['Sheet1']
        source_ws = source_wb['Sheet1']
        for row in source_ws.iter_rows():
            for cell in row:
                target_ws[cell.coordinate].font = cell.font.copy()
                target_ws[cell.coordinate].border = cell.border.copy()
                target_ws[cell.coordinate].fill = cell.fill.copy()
                target_ws[cell.coordinate].number_format = cell.number_format
                target_ws[cell.coordinate].alignment = cell.alignment.copy()

alpha = load_workbook(f'{file_path}.xlsx')
copy_formatting(load_workbook('formatting_template.xlsx'),alpha)

print("Numberising...")
for row in alpha['Sheet1'].iter_rows():
    for cell in row:
        if cell.value == None:
            continue
        elif cell.value.startswith('Unnamed'):
            cell.value = None
            continue
        else:
            for letter, number in code.items():
                cell.value = cell.value.replace(letter,str(number))
                cell.font = Font(name='Calibri (Body)',
                                 size=14,
                                 color = '000000')


print("Saving...")
alpha.save(f'new_sudoku{file_path[10]}{file_path[-2:]}.xlsx')

print("Success!")

# # except:
# #     print("Some kinda problem happened")

